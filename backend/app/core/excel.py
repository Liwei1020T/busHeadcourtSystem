"""
Excel parsing helpers used by upload endpoints.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
from typing import Any, Dict, Iterable, Optional, Sequence, Tuple

from openpyxl import load_workbook
import re


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = text.replace("\t", " ")
    while "  " in text:
        text = text.replace("  ", " ")
    return text


def _canonical_header(value: Any) -> str:
    normalized = _normalize_header(value)
    return normalized.replace(" ", "").replace("_", "").replace("-", "")


@dataclass(frozen=True)
class ExcelRow:
    row_number: int
    values: Dict[str, Any]


@dataclass(frozen=True)
class ExcelTable:
    sheet_name: str
    header_row_number: int
    headers: Sequence[str]
    rows: list[ExcelRow]


def _find_header_row(
    ws,
    must_include: set[str],
    prefer_include: set[str],
    max_scan_rows: int = 50,
    min_prefer_matches: int = 0,
) -> Optional[Tuple[int, Sequence[str], int]]:
    best: Optional[Tuple[int, Sequence[str], int]] = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        if not row:
            continue
        headers = [_canonical_header(cell) for cell in row]
        header_set = {h for h in headers if h}

        if not must_include.issubset(header_set):
            continue

        prefer_matches = len(header_set.intersection(prefer_include))
        if prefer_matches < min_prefer_matches:
            continue
        score = prefer_matches * 1000 + len(header_set)
        if best is None or score > best[2]:
            best = (row_idx, headers, score)

    return best


def _is_non_empty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _count_valid_sample_rows(ws, header_row_idx: int, headers: Sequence[str], required_non_empty: set[str], sample_size: int) -> int:
    required_indices = [idx for idx, header in enumerate(headers) if header in required_non_empty]
    if not required_indices:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + sample_size, values_only=True):
        if not row:
            continue
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        if all((idx < len(row) and _is_non_empty(row[idx])) for idx in required_indices):
            count += 1
    return count


def read_rows_from_best_sheet(
    xlsx_bytes: bytes,
    must_include: set[str],
    prefer_include: Optional[set[str]] = None,
    sheet_name_exclude_prefixes: Optional[Sequence[str]] = None,
    min_prefer_matches: int = 0,
    required_non_empty_in_sample: Optional[set[str]] = None,
    min_valid_sample_rows: int = 1,
    sample_size: int = 15,
) -> list[ExcelRow]:
    """
    Read rows from the best-matching worksheet in an XLSX.

    The selected sheet is the one that contains a header row matching `must_include`
    (case-insensitive after canonicalization), scoring higher for additional matches
    in `prefer_include`.
    """
    prefer_include = prefer_include or set()
    sheet_name_exclude_prefixes = sheet_name_exclude_prefixes or ("note",)

    wb = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)

    best_sheet = None
    best_header = None
    best_sample_valid = 0

    for ws in wb.worksheets:
        title = (ws.title or "").strip().lower()
        if any(title.startswith(prefix) for prefix in sheet_name_exclude_prefixes):
            continue

        header = _find_header_row(
            ws,
            must_include=must_include,
            prefer_include=prefer_include,
            min_prefer_matches=min_prefer_matches,
        )
        if header is None:
            continue

        header_row_idx, headers, score = header
        required_non_empty_in_sample = required_non_empty_in_sample or set()
        sample_valid = _count_valid_sample_rows(ws, header_row_idx, headers, required_non_empty_in_sample, sample_size) if required_non_empty_in_sample else sample_size

        if required_non_empty_in_sample and sample_valid < min_valid_sample_rows:
            continue

        score_with_sample = score + (sample_valid * 100)
        if best_header is None or score_with_sample > best_header[2]:
            best_sheet = ws
            best_header = (header_row_idx, headers, score_with_sample)
            best_sample_valid = sample_valid

    if best_sheet is None or best_header is None:
        return []

    header_row_idx, headers, _score = best_header
    rows: list[ExcelRow] = []
    for row_idx, row in enumerate(best_sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        values: Dict[str, Any] = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            cell_value = row[col_index] if col_index < len(row) else None
            values[header] = cell_value
        rows.append(ExcelRow(row_number=row_idx, values=values))

    return rows


def read_table_from_best_sheet(
    xlsx_bytes: bytes,
    must_include: set[str],
    prefer_include: Optional[set[str]] = None,
    sheet_name_exclude_prefixes: Optional[Sequence[str]] = None,
    min_prefer_matches: int = 0,
    required_non_empty_in_sample: Optional[set[str]] = None,
    min_valid_sample_rows: int = 1,
    sample_size: int = 15,
) -> Optional[ExcelTable]:
    wb = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
    prefer_include = prefer_include or set()
    sheet_name_exclude_prefixes = sheet_name_exclude_prefixes or ("note",)

    best_sheet = None
    best_header = None

    for ws in wb.worksheets:
        title = (ws.title or "").strip().lower()
        if any(title.startswith(prefix) for prefix in sheet_name_exclude_prefixes):
            continue

        header = _find_header_row(
            ws,
            must_include=must_include,
            prefer_include=prefer_include,
            min_prefer_matches=min_prefer_matches,
        )
        if header is None:
            continue

        header_row_idx, headers, score = header
        sample_valid = _count_valid_sample_rows(ws, header_row_idx, headers, required_non_empty_in_sample or set(), sample_size) if required_non_empty_in_sample else sample_size
        if required_non_empty_in_sample and sample_valid < min_valid_sample_rows:
            continue

        score_with_sample = score + (sample_valid * 100)
        if best_header is None or score_with_sample > best_header[2]:
            best_sheet = ws
            best_header = (header_row_idx, headers, score_with_sample)

    if best_sheet is None or best_header is None:
        return None

    header_row_idx, headers, _score = best_header
    rows: list[ExcelRow] = []
    for row_idx, row in enumerate(best_sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        values: Dict[str, Any] = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            cell_value = row[col_index] if col_index < len(row) else None
            values[header] = cell_value
        rows.append(ExcelRow(row_number=row_idx, values=values))

    return ExcelTable(sheet_name=best_sheet.title, header_row_number=header_row_idx, headers=headers, rows=rows)


def read_first_sheet_rows(xlsx_bytes: bytes) -> Iterable[ExcelRow]:
    """
    Backwards-compatible wrapper: prefer the best sheet containing at least `personid`.
    """
    return read_rows_from_best_sheet(
        xlsx_bytes,
        must_include={"personid"},
        prefer_include={"name"},
        required_non_empty_in_sample={"personid"},
        min_valid_sample_rows=1,
    )


def coerce_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def coerce_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def coerce_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Common exports include weekday prefixes/suffixes, e.g. "12/01/2026 (Mon)" or "Tue-24/12/2024".
    text = re.sub(r"\s*\(.*\)\s*$", "", text)
    text = re.sub(r"^[A-Za-z]{3,9}-", "", text)
    text = text.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def coerce_time(value: Any) -> Optional[time]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    text = str(value).strip()
    if not text or text == ".":
        return None
    text = text.upper()
    for fmt in ("%I:%M %p", "%I:%M%p", "%I:%M:%S %p", "%I:%M:%S%p", "%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def build_scanned_at(scanned_on: date, shift: str) -> datetime:
    """
    Build a deterministic timestamp for a given scanned_on date when the source only provides a date.
    """
    if shift == "morning":
        t = time(8, 0)
    elif shift == "night":
        t = time(18, 0)
    else:
        t = time(12, 0)
    return datetime.combine(scanned_on, t)
